import scipy.signal as signal
import openpyxl as xl
import csv
import numpy as np
import matplotlib.pyplot as plt
import os as os
from scipy.stats import linregress


# return first numsamples samples of data
def sample_data(optical_data_folder,
                force_data_folder,
                num_samples):
    PLOT = False

    opt_and_force_data_array = {}




    folder_path = os.getcwd() + "\\" + optical_data_folder
    optical_data_filenames = os.listdir(folder_path)

    for fileno in range(0, len(optical_data_filenames)):
        key = "Post {0}".format(fileno)
        opt_samples, opt_samples_values = Sample_Excel_Data(optical_data_folder,
                                                           fileno,
                                                           num_samples,
                                                           True,
                                                           id=key,
                                                           scale = 1,
                                                           Filter=False,
                                                           sampling_freq=6.667,
                                                           cutoff=5,
                                                           peak_height=(1.5, 15),
                                                           shift=-15)

        force_samples, force_samples_values = Sample_Excel_Data(force_data_folder,
                                                               fileno,
                                                                num_samples,
                                                               True,
                                                               id=key,
                                                               scale = 1,
                                                               Filter=True,
                                                               sampling_freq=50,
                                                               cutoff=2.5,
                                                               peak_height=(.001, .007),
                                                               shift=-35)

        opt_and_force_data_array.update({key: {'opt_samples_locations': opt_samples,
                                               'opt_samples_values': opt_samples_values,
                                               'force_samples_locations': force_samples,
                                               'force_samples_values': force_samples_values}})

    return opt_and_force_data_array


def Sample_Excel_Data(folder_name,
                      filenum,
                      numsamples,
                      plot_time_output = True,
                      id="0",
                      scale=1,
                      Filter=False,
                      sampling_freq=50,
                      cutoff=5,
                      peak_height = (2, 15),
                      shift=0):

    folder_path = os.getcwd() + "\\" + folder_name
    print(folder_path)
    filenames = sorted(os.listdir(folder_path))
    print(filenames)

    wb = xl.load_workbook(folder_path + "\\" + filenames[filenum])
    ws = wb.active

    # A1
    if "X000" in filenames[filenum]:
        xcol = ws['C']
    else:
        xcol = ws['B']

    x = []

    for timepoint in range(4, len(xcol)):
        x.append(xcol[timepoint].value)

    xfilt = np.array(x)

    if Filter:
        b, a = signal.butter(4, cutoff, 'low', fs=sampling_freq)
        xfilt = -signal.filtfilt(b, a, x)
        b, a = signal.butter(4, 10, 'low', fs=sampling_freq)
        x = -signal.filtfilt(b, a, x)



    peaks_optical, _ = signal.find_peaks(np.diff(xfilt), height=peak_height, distance = int(sampling_freq))
    opt_samples = peaks_optical + shift
    opt_samples_values = np.array(x)[opt_samples] / scale

    if plot_time_output:
        plt.grid(True)
        print(id)
        plt.plot(np.diff(xfilt))
        plt.plot(peaks_optical, np.diff(np.array(xfilt))[peaks_optical], 'x')
        plt.show()

        plt.grid(True)
        plt.plot(.02 * np.arange(len(x)), np.array(x) / scale)
        plt.plot(.02 * opt_samples, opt_samples_values, 'x')
        plt.xlabel("time elapsed (s)")
        plt.ylabel("force sensor output (V)")
        plt.show()
    return opt_samples[0:numsamples], opt_samples_values[0:numsamples]

leg = []
for f in range(0, 6):
    leg.append(f)
    force_sample_indices, force_sample_values = Sample_Excel_Data(r"03 01 2023/1x_SilasticQ74840_Contact_Below_Cap_Force",
                                                               f,
                                                               30,
                                                               plot_time_output=False,
                                                               scale = 1,
                                                               Filter=True,
                                                               sampling_freq=50,
                                                               cutoff=2.5,
                                                               peak_height=(.001, .03),
                                                               shift=-15)

# FvO = sample_data("08232022_SilasticQ74840_Optical", "08232022_SilasticQ74840_Force", 19)
#
#
# for post in FvO.keys():
#     plt.plot(FvO[post]["opt_samples_values"], FvO[post]["force_samples_values"])
#
#     result = linregress(FvO[post]["opt_samples_values"], FvO[post]["force_samples_values"])
#     print()
#     print(" R Sq.")
#     print(result.rvalue ** 2)
#     print(" Slope ")
#     print(result.slope)
#     plt.plot(FvO[post]["opt_samples_values"], result.slope * FvO[post]["opt_samples_values"] + result.intercept)


    plt.plot(np.arange(30)*.1, force_sample_values)

    result = linregress(np.arange(30)*.1, force_sample_values)
    print()
    print(" R Sq.")
    print(result.rvalue ** 2)
    print(" Slope ")
    print(result.slope)
    plt.plot(np.arange(30)*.1, result.slope*np.arange(30)*.1 + result.intercept, ".")

    plt.grid(True)
    # plt.legend(leg)
    plt.legend(["Measured", "Linear fit"])
    plt.title("Force vs Displacement")
    plt.xlabel("Unscaled X Position (Pixels)")
    plt.ylabel("Unscaled Force (V)")
    plt.show()