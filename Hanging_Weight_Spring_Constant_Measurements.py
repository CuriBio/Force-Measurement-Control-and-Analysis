import openpyxl as xl
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import linregress
from scipy.stats import shapiro

data_columns = ['A', 'C']#, 'E', 'G', 'I', 'K']
weight_num_12x = np.arange(6)
weight_12x = .000050 * 9.8  # kg * m/s sq
scale = 230 * 1000  # pixels / m
num_posts = 6
num_cycles = 1
num_weights = len(weight_num_12x)
force = weight_num_12x * weight_12x
spring_constants_12x = np.zeros((len(data_columns), num_posts))

spreadsheet_path = r"C:\Users\kevin\Downloads\RowC&DPostMeasurements\SpreadSheets\wb_fin.xlsx"
# v data from analysis detailed in https://app.asana.com/0/search?q=inertia&sort=completion_time&child=1206099065617078
# spreadsheet_path = r"G:\Shared drives\IR&D Projects\Mantarray-V1\3_Technical Development\MAv1_Field-to-Force\Position To Force\Spring Constant Estimates -- Hanging Weight\spring_constant_test_11252023\weight_hanging_test_set_1.xlsx"
print(spreadsheet_path)

wb = xl.load_workbook(spreadsheet_path, data_only=True)
ws = wb.active
for mold_num, mold in enumerate(data_columns):
    print("mold", mold_num + 1)
    print()
    raw_col = np.array(ws[mold])
    raw_disp_values_list = []

    # load measurements in spreadsheet into list
    for cell in range(0, len(raw_col)):
        if type(raw_col[cell].value) == float or type(raw_col[cell].value) == int:
            raw_disp_values_list.append(raw_col[cell].value)

    # absolute pixel position of post head, convert list to numpy array
    position_values = np.array(raw_disp_values_list).reshape(num_posts, num_cycles, num_weights) / scale
    # change in position from unweighted position
    displacement_values = (position_values.T - position_values[:, :, 0].T).T
    # mean change from unweighted position
    mean_disp_values = np.mean(displacement_values, axis=1)
    # standard deviation of change from unweighted position
    stdev_disp_values = np.std(displacement_values, axis=1)

    # Example position - to - force
    for post in range(0, num_posts):

        # plot mean displacements from all loading cycles for a given post
        # plt.errorbar(
        #     force,
        #     mean_disp_values[post],
        #     stdev_disp_values[post],
        #     marker='s',
        #     ms=5,
        #     capsize=10,
        #     fmt='_'
        # )



        result = linregress(force, displacement_values[post, 0])

        print("post", post + 1)
        print(" R Sq.")
        print(result.rvalue**2)
        print(" Stiffness ")
        print(result.slope**-1)
        print()

        # plot displacements from one loading cycle
        plt.plot(
            force,
            displacement_values[post, 0],
            "D"
        )
        plt.plot(force,
                 result.slope*force + result.intercept)

    plt.ylabel("post head displacement (m)")
    plt.xlabel("applied force (N)")
    # plt.title("regression of {0} loading cycles of post {1}".format(num_cycles, post + 1))
    plt.title("regression of 1 loading cycle of post {0}".format(post + 1))

    plt.grid(True)
    plt.tight_layout()
    plt.show()


    # Plot all regression slopes
    for post in range(0, num_posts):
        result = linregress(force, mean_disp_values[post]) # mean displacement of all cycles
        # result = linregress(force, displacement_values[post, 0]) # one cycle
        # print()
        # print("post" + str(post + 1))
        # print(" R Sq.")
        # print(result.rvalue ** 2)
        # print(" Stiffness ")
        # print(result.slope ** -1)
        spring_constants_12x[mold_num, post] = (result.slope ** -1)

    k_mean = np.mean(spring_constants_12x[mold_num])
    k_2stdev = 2*np.std(spring_constants_12x[mold_num])
    print()
    print("mean stiffness:")
    print(str(k_mean) + "N/m")
    print("2x standard deviation stiffness")
    print(str(k_2stdev) + "N/m")
    print("as a percentage of mean stiffness")
    print(str(k_2stdev/k_mean * 100) + "%")

    print("Shapiro-Wilk normality test p-value")
    stat, p = shapiro(spring_constants_12x[mold_num])
    print(p)
    print()

bar_positions = np.arange(len(data_columns))
mean_k = np.mean(spring_constants_12x, axis=1)  # N/m
x2stdev_k = 2*np.std(spring_constants_12x, axis=1)  # N/m
x2sterr_k = x2stdev_k/np.sqrt(num_posts)
# plt.margins(.005)
plt.bar(bar_positions,
             mean_k, # mean for each mold
             yerr=x2sterr_k,
             capsize=10,
        tick_label=bar_positions,
        edgecolor="black",
        facecolor="gainsboro",
        color="white",
        width=.75)
# plt.xticks(bar_positions, labels=bar_positions + 1)
plt.xticks(bar_positions, labels=["D", "C"])

for bar_position in bar_positions:
    plt.plot(bar_position, .75, marker=f"${np.around((mean_k[bar_position]), 2)}N/m$", color="black", markersize=40)
    # plt.plot(bar_position, .625, marker=f"$+/-{np.around((x2stdev_k[bar_position]), 2)}N/m$", color="black", markersize=60)
    # plt.plot(bar_position, .5, marker=f"$+/-{np.around((x2stdev_k[bar_position]/mean_k[bar_position])*100, 1)} \%$", color="black", markersize=60)

    for k in range(0, num_posts):
        plt.plot(bar_position
                 + .6*np.random.rand()
                 - .3,
                 # spring_constants_12x[bar_position, k], marker="$post {0}$".format(k + 1), color="black", markersize=20)
                 spring_constants_12x[bar_position, k], marker="D", color="black", markersize=5)

plt.plot(np.arange(-1, len(bar_positions) + 1), 1.98*np.ones(len(bar_positions) + 2), color="blue")
plt.plot(-.75, 2.04, marker="$1.98 N/m$", markersize=40, color="blue")

plt.title("spring constants from regression of 1 loading cycle")
# plt.title("spring constants from regression of 4 loading cycles")

plt.xlabel("row")
plt.ylabel("stiffness (N/m)")
plt.grid(True)
plt.show()

# # relate post diameter to spring constant
# post_diameters = []
# dia_col = np.array(ws['C'])
# for cell in range(0, num_posts):
#     cell_of_interest = cell * num_cycles * num_weights
#     if type(dia_col[cell_of_interest].value) == float or type(dia_col[cell_of_interest].value) == int:
#         post_diameters.append(dia_col[cell_of_interest].value/scale)
#
# for k in range(0, len(spring_constants_12x)):
#     plt.scatter(
#         spring_constants_12x[k],
#         np.pi*(post_diameters[k])**4/64,
#         marker="$post {0}$".format(k + 1),
#         s=500,
#         c="black"
#     )
# plt.xlabel("post spring constant (N/m)")
# plt.ylabel("post cross sectional moment of inertia ($m^4$)")
# plt.grid(True)
# plt.show()
# print(post_diameters)

